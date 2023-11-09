import torch.optim
from torch.utils.tensorboard import SummaryWriter
import os
from argparse import ArgumentParser
import numpy as np
from tqdm.autonotebook import tqdm
from src import  summary_utils, utils, mymodels
from place2.places2 import Places2
from place2.val_places2 import Places2_yuanchicun
from torch.utils.data import DataLoader
from torchvision.utils import save_image
import time
import torch.nn as nn
import sys
import xlwt
from openpyxl import load_workbook

#记得修改网络输出的通道数，改为3

print('torch.cuda.device_count()',torch.cuda.device_count())
# exit()
utils.seed(num=123,deterministic=True)
if torch.cuda.device_count() ==5:
    device = torch.device('cuda:4' if torch.cuda.is_available() else 'cpu')
elif torch.cuda.device_count()==4:
    device = torch.device('cuda:3' if torch.cuda.is_available() else 'cpu')
elif torch.cuda.device_count()==3:
    device = torch.device('cuda:2' if torch.cuda.is_available() else 'cpu')
else:
    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    
print('device',device)
# exit(0)
def main(args):
    start_time=time.time()
    save_root_txt='{:s}/ckpt/{:s}/{:s}/set.txt'.format(args.save_dir, args.shutter,args.interp)
    fo = open(save_root_txt, 'a')
    fo.write('epoch:{}, '.format(args.max_epochs))
    fo.write('batch_size:{}, '.format(args.batch_size))
    fo.write('cfa_size:{}, '.format(args.cfa_size))
    fo.write('slr:{}, '.format(args.slr))
    fo.write('mlr:{}, '.format(args.mlr))
    fo.write('shutter:{}, \n'.format(args.shutter))
    fo.write('interpolation:{}, '.format(args.interp))
    fo.write('decoder:{}, '.format(args.decoder))
    fo.write('proxy function:{}, \n'.format(args.init))
    fo.write('bool_noise:{}, '.format(args.bool_noise))
    fo.write('noise_std:{}, '.format(args.noise_std))
    fo.write('loss:{}, '.format(args.loss))
    fo.write('vis_interval:{}, \n'.format(args.vis_interval))
    fo.write('w_zhi:{}, '.format(args.w_zhi))
    fo.write('alpha:{}, '.format(args.alpha))
    fo.write('gaussian_weight_size:{}, \n'.format(args.gaussian_weight_size))
    fo.close()
    # torch.cuda.empty_cache() #清除pytorch缓存
    # if 'lsvpe' or 'lrgbw' not in args.shutter: ##注意 or 和in一块使用，in 的运算符顺序优于or，此结果为lsvpe，达不到想要的结果
        # print(args.shutter) #rgb
        # print('lsvpe' or 'lrgbw' not in args.shutter) #lsvpe
        #我们想要的结果是false，实际结果并不是，因为or 和 in的运算顺序问题，in 大于or
        # if 'lsvpe'' not in args.shutter  or 'lrgbw' not in args.shutter: 这样才可以达到效果
    # if 'lrgbw' not in args.shutter:
    #     print('SETTING MLR TO 2E-4 since we are only training decoder') #将MLR设置为2E-4，因为我们只是训练解码器
    #     args.mlr = 2e-4
    args.slr = float(args.slr) #slr--shutter lr
    args.mlr = float(args.mlr) #mlr--model lr
    # print('args.mlr',args.mlr) # 0.0002
    # exit()

    if not os.path.exists(args.save_dir):
        os.makedirs('{:s}/images'.format(args.save_dir))
        os.makedirs('{:s}/ckpt'.format(args.save_dir))

    if not os.path.exists(args.log_dir):
        os.makedirs(args.log_dir)
    
    shutter=mymodels.define_myshutter(args.shutter,args)
    # print('shutter',shutter) #LSVPE() RGBW() LRGBW()
    decoder=mymodels.define_mydecoder(args.decoder,args)
    model=mymodels.define_mymodel(shutter,decoder,args,get_coded=False)
    # print('model',model)
    # model.cuda() #原版
    """
    #查看有哪些需要学习的参数
    # print(model.parameters())
    # for name ,parameters in model.shutter.named_parameters():
        # print(name,':',parameters.size())
    
    for p in model.shutter.named_parameters():
        print('parameters',p) #输出 有哪些可学习参数

    """
    mylogs_dir='{:s}/{:s}/{:s}/{:s}/'.format(args.log_dir,args.shutter,args.interp,args.init)
    print(args.log_dir)
    writer = SummaryWriter(log_dir=mylogs_dir)
    optim = utils.mydefine_optim(model, args)
    # dataset_train=Places2(img_root=args.root,block_size=args.block_size,std=args.noise_std,bool_noise=args.bool_noise,split='train') #
    dataset_train = Places2(img_root=args.root, block_size=args.block_size, std=args.noise_std,
                            bool_noise=args.bool_noise, split='train_val_batch_128')
    # dataset_train=Places2(img_root=args.root,block_size=args.block_size,split='train_batch') #输入为随机剪切的
    print('len(dataset_train)',len(dataset_train))
    dataset_val=Places2_yuanchicun(img_root=args.root, std=args.noise_std,
                        bool_noise=args.bool_noise,split='val') #
    # dataset_val=Places2_yuanchicun(img_root=args.root, std=args.noise_std,bool_noise=args.bool_noise,split='val') #
    print('len(dataset_val)',len(dataset_val))
    # exit()
    train_dataloader=DataLoader(dataset_train,
                      batch_size=args.batch_size,
                      num_workers=args.num_workers,
                      shuffle=False,
                      )
    val_dataloader=DataLoader(dataset_val,
                      batch_size=1,
                      num_workers=args.num_workers,
                      shuffle=False,
                    )

    loss_fn = utils.define_loss(args)
    # print('loss_fn',loss_fn)
    best_val_rgb_psnr = 0
    psnr_rgb_mean_list=[]
    print('len(train_dataloader)',len(train_dataloader)) #48 注意dataloader的长度和batch_size有关系，长度等于len(datasets_train)/batch_size向上取整
    print('len(val_dataloader)',len(val_dataloader)) #48
    # if torch.cuda.device_count() > 1:  # 检查电脑是否有多块GPU
        # print(f"Let's use {torch.cuda.device_count()} GPUs!")
        # model = nn.DataParallel(model)  # 将模型对象转变为多GPU并行运算的模型
    model.to(device)
    snapshot = '{:s}/ckpt/{:s}/{:s}/{:s}/{:s}_epoch.pth'.format(args.snapshot, args.shutter, args.interp,args.init, args.test_epoch)
    continue_epoch=0
    if args.resume==True:
        print("从检查点处开始训练")
        print('snapshot',snapshot)
        checkpoint = torch.load(snapshot)
        try:
            model.load_state_dict(checkpoint['model_state_dict'])
        except KeyError:
            model.load_state_dict(checkpoint)
        continue_epoch=args.continue_epoch
    # print('model.shutter.parameters',model.shutter.parameters)
    # print('model.decoder.parameters',model.decoder.parameters)
    total_steps = 1 #原版为0
    for j in tqdm(range(continue_epoch,args.max_epochs)):
        model.train()
        for  step,(model_input, gt) in enumerate(tqdm(train_dataloader)):
            # print('step',step)
            # print('model_input.shape',model_input.shape) #orch.Size([1, 4, 186, 317])
            # print('gt.shape',gt.shape) #torch.Size([1, 4, 186, 317])
            model_input = model_input.to(device)
            gt = gt.to(device)
            gt_rgb=gt[:,:3,:,:]
            # print(gt_rgb.shape) #torch.Size([8, 3, 256, 256])
            # exit()
            # optim.zero_grad(set_to_none=True)#原版

            restored,cfa_img= model(model_input,train=True)#修改版本
            #实时监控cfa变化
            save_image(cfa_img[:,:args.cfa_size[0],:args.cfa_size[1]],'learncfa.png')
            # print(restored.shape) #torch.Size([8, 3, 256, 256])
            # print(cfa_img.shape) #torch.Size([3, 256, 256])
            #cfa_img为3通道RGB tensor直接可以保存，维度为[3,h,w]

            # restored,end_params= model(model_input,train=True)#测试梯度版本
            # print('restored.shape',restored.shape)
            # print('restored',restored)
            # print(restored>1)
            train_loss = loss_fn(restored, gt_rgb)
            print('train_loss',train_loss)
            optim.zero_grad()
            train_loss.backward()
            # print('image.grad',image.grad)
            # print('image.grad',image.grad==0.0)
            # print('image.requires_grad',image.requires_grad)
            # print('image.is_leaf', image.is_leaf)
            # print('end_params.requires_grad',end_params.requires_grad)
            # print('end_params.is_leaf',end_params.is_leaf)
            # sys.exit()
            optim.step()
            """
            if ((total_steps ) % args.save_loss == 0):
                writer.add_scalar('train_loss', train_loss, total_steps )
                print('保存训练损失')
            total_steps += 1
            """
        # """
        if (j+ 1) % args.save_model_interval == 0 or (j + 1) == args.max_epochs:
            save_dir='{:s}/ckpt/{:s}/{:s}/{:s}/{:d}_{:d}_epoch.pth'.format(args.save_dir, args.shutter,args.interp,args.init,args.code,j+1)
            utils.save_mychkpt(model,optim,save_dir) #本篇论文自带
            # utils.save_myckpt(save_dir, [('model', model)], [('optimizer', optim)], j+1)  # PConv
        # """
        if (j+1) % args.save_cfa_epoch==0:
            save_cfa_dir='{:s}/{:s}/{:s}/{:s}/{:d}_{:d}_epoch.png'.format(args.save_cfa_root,args.shutter,args.interp,args.init,args.code,j+1)
            for i in range(args.cfa_size[0]):
                for k in range(args.cfa_size[1]):
                    if cfa_img[0, i, k] == 0 and cfa_img[1, i, k] == 0 and cfa_img[2, i, k] == 0:
                        cfa_img[:3, i, k] = 1
            save_image(cfa_img[:,:args.cfa_size[0],:args.cfa_size[1]],save_cfa_dir)
            # exit(0)
        if (j + 1) % args.vis_interval == 0:
            if val_dataloader is not None:
                with torch.no_grad():
                    model.eval()
                    val_psnrs_rgb = []
                    for (model_input, gt) in tqdm(val_dataloader):
                        # print('model_input',model_input)
                        model_input = model_input.to(device)
                        gt = gt.to(device)
                        gt_rgb=gt[:,:3,:,:]
                        restored,cfa_img = model(model_input,train=False)#修改版本
                        psnr_rgb=summary_utils.get_psnr(restored,gt_rgb)
                        # print('psnr_rgb',psnr_rgb)
                        save_dir='{:s}/images/{:s}/{:s}/{:s}/{:d}_{:d}_epoch.pth'.format(args.save_dir, args.shutter,args.interp,args.init,j+1,args.code)
                        # save_image(restored_rgb,dir)
                        val_psnrs_rgb.append(psnr_rgb)

                    psnr_rgb_mean_list.append(np.mean(val_psnrs_rgb))
                    print('np.mean(val_psnrs_rgb)',np.mean(val_psnrs_rgb))
                    fo = open(save_root_txt, 'a')
                    fo.write('psnr_rgb_mean_list:\n{} {}\n'.format(psnr_rgb_mean_list, len(psnr_rgb_mean_list)))
                    fo.close()
                    if np.mean(val_psnrs_rgb) > best_val_rgb_psnr:
                                print(f'BEST PSNR: '
                                    #   f'{np.mean(val_psnrs)}, SSIM: {np.mean(val_ssims)}, LPIPS: {np.mean(val_lpips)}')
                                      f'{np.mean(val_psnrs_rgb)}, SSIM: {0}, LPIPS: {0}')
                                best_val_rgb_psnr = np.mean(val_psnrs_rgb)
                                save_dir='{:s}/ckpt/{:s}/{:s}/{:s}/{:d}_best_epoch.pth'.format(args.save_dir, args.shutter,args.interp,args.init,args.code)
                                utils.save_mychkpt(model,optim,save_dir)  #本篇论文自带
                                # utils.save_myckpt(save_dir,[('model', model)], [('optimizer', optim)], 'best') #PConv
                                # save_dir='{:s}/images/{:s}/{:s}/{:s}/best_epoch.pth'.format(args.save_dir, args.shutter,args.interp,args.init)
                                # save_image(restored_rgb,x)

                                # utils.save_myckpt('{:s}/ckpt_rgbw/scatter/best.pth'.format(args.save_dir),
                                    #  [('model', model)], [('optimizer', optim)], j + 1)
                                # utils.save_chkpt(model, optim, checkpoints_dir, epoch=epoch, best=True)
    print('psnr_rgb_mean_list',psnr_rgb_mean_list,'--len(psnr_rgb_mean_list)',len(psnr_rgb_mean_list))
    print('best_val_rgb_psnr',best_val_rgb_psnr)

    writer.close()
    print("epochs:  ",args.max_epochs)
    print("batch_size:",args.batch_size)
    print("cfa_size:",args.cfa_size)
    print("slr:",args.slr)
    print("mlr:",args.mlr)
    print("shutter: ",args.shutter ) 
    print("interpolation: ",args.interp)
    print("decoder:",args.decoder)
    print('代理函数：',args.init)
    print('固定cfaw初始化值为：',args.w_zhi)
    print("alpha: ",args.alpha)
    print("gaussian_weight_size: ",args.gaussian_weight_size)
    print("是否加噪声:",args.bool_noise)
    print("噪声等级std:",args.noise_std)
    end_time=time.time()
    print(f"程序运行时间：{end_time-start_time:.4f}s")
    fo = open(save_root_txt, 'a')
    fo.write('程序运行时间:{:.4f}s \n'.format(end_time-start_time))
    fo.write('psnr_rgb_mean_list:\n{} {}\n'.format(psnr_rgb_mean_list,len(psnr_rgb_mean_list)))
    fo.write('best_val_rgb_psnr:{}\n\n'.format(best_val_rgb_psnr))
    fo.close()
    return best_val_rgb_psnr
    
    

if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--root', type=str, default='../canon_data/Gehler_Shi568_Convert_Qiao')
    # parser.add_argument('--root', type=str, default='../canon_data/568_great_test') #568_great_test  568_great
    # parser.add_argument('--root', type=str, default='../canon_data/data_tiaocan')
    parser.add_argument('--snapshot', type=str,default='../snapshots/default')
    parser.add_argument('--test_epoch', type=str, default='Big36_best')
    parser.add_argument('--continue_epoch', type=int, default=20) #如果从模型中间开始训练必须修改从第几个epoch开始训练
    parser.add_argument('--resume',
                    type=bool,
                    default=False,
                    help='是否需要从检查点处开始训练，如需要请修改，test_epoch,并且其他参数也务必相同')
    parser.add_argument('-b', '--block_size',
                    help='delimited list input for block size in format(格式中块大小的分隔列表输入) %,%,%',
                    default=[4,128,128]) #default=[4,324,487] default='8,512,512' 8,1359,2041
    # parser.add_argument('--reg', type=float, default=100.0, help='regularization on lpips loss(lpips损失的正规化)')
    parser.add_argument('--cfa_size', help='cfa的尺寸大小,只针对lrgbw有效%,%,%',
                    default=[4,4])
    parser.add_argument('--k', type=int, default=3) #原版default=3
    parser.add_argument('--p', type=float, default=1.0)
    parser.add_argument('--bool_noise',type=bool,default=True)
    parser.add_argument('--noise_std',type=float,default=0.01)
    parser.add_argument('--max_epochs', type=int, default=20) #原版6000
    parser.add_argument('--mlr', help='model_lr', type=str, default='2e-4') #原版 default='5e-4'
    parser.add_argument('--slr', help='shutter_lr', type=str, default='1e-3') #原版 default='2e-4'
    parser.add_argument('--batch_size', type=int,default=8) #原版default=2
    parser.add_argument('--w_zhi', type=float,default=0.001)
    parser.add_argument('--alpha', type=float,default=2.5)
    parser.add_argument('--gaussian_weight_size', type=int,default=5)
    parser.add_argument('--num_workers', type=int, default=4) #原版1
    parser.add_argument('--interp', type=str, choices=['none', 'bilinear', 'scatter','pconv','gaussian'],
                        default='gaussian') #原版required=True, default='pconv'
    parser.add_argument('--init', type=str, choices=['softmax_tau', 'gumbel_softmax_tau_1', 'gumbel_softmax_tau'], 
                        default='softmax_tau',
                help='针对可学习RGBW CFA的形式,反向传播代理函数问题,'
                        'softmax_tau:带温度系数的softmax函数,gama=2.5e-5,可去函数里面调整,tau根据迭代次数进行变化,'
                        'gumbel_soft_tau_1:Gumbel_Softmax函数,tau为初始值1,'
                        'gumbel_softmax_tau:Gumbel_Softmax,tau可变,根据迭代次数变化')
    parser.add_argument('--loss', type=str, choices=['mpr', 'l1', 'l2_lpips', 'l2'], default='l2') #原版 default='l2_lpips'
    parser.add_argument('--decoder', type=str,
                        choices=['unet', 'mpr', 'dncnn', 'myunet','sgnet',
                                 'mysgnet','drunet','rlfn','nafnet',
                                 'mysgnet1','rrdbnet','rdunet'],
                        default='mysgnet1') #原版default='unet'
    parser.add_argument('--shutter', type=str, default='lrgbw') #quad 原版required=True, default='lsvpe'  default='rgbw'
    parser.add_argument('--save_dir', type=str, default='../snapshots/default')
    parser.add_argument('--log_dir', type=str, default='../mylogs')
    parser.add_argument('--save_model_interval', type=int, default=20) #保存模型时间间隔default=50000
    parser.add_argument('--vis_interval', type=int, default=1)#垂直间隔,每几个epoch去验证集
    parser.add_argument('--save_loss', type=int, default=10000000000000000)#经过几个迭代保存一次loss
    parser.add_argument('--save_cfa_epoch', type=int, default=1)#经过几个epoch保存一次cfa图像
    parser.add_argument('--save_cfa_root', type=str, default='../cfa_image')
    parser.add_argument('--code', type=int, default=1)
    args = parser.parse_args()
    # main(args)


    # """
    read_root_txt='{:s}/ckpt/{:s}/{:s}/set_canshu.txt'.format(args.save_dir, args.shutter,args.interp)
    save_root_xlsx='{:s}/ckpt/{:s}/{:s}/demo.xlsx'.format(args.save_dir, args.shutter,args.interp)
    #加载excel，注意路径要与脚本一致
    wb = load_workbook(save_root_xlsx)
    #激活excel表
    sheet = wb.active
    fo_main=open(read_root_txt,'r')
    read_lines=fo_main.readlines()
    print('read_lines',read_lines)
    print('len(read_lines)',len(read_lines))
    code=0
    for m  in tqdm(range(1,len(read_lines))):
    # print('m',m)
    # exit()
        cfa_size=[]
        parameter=read_lines[m].strip('\n').split(',')
        # print(parameter)
        # exit()
        args.code=int(parameter[0])
        args.batch_size=int(parameter[1])
        args.slr=parameter[2]
        args.alpha=float(parameter[3])
        args.w_zhi=float(parameter[4])
        args.gaussian_weight_size=int(parameter[5])
        args.mlr=parameter[6]
        cfa_sz=parameter[7].split('x')
        cfa_size.append(int(cfa_sz[0]))
        cfa_size.append(int(cfa_sz[1]))
        args.cfa_size=cfa_size
        # print('args.cfa_size',args.cfa_size)
        # exit()
        # best_rgb_psnr,best_w_psnr=1,1
        best_rgb_psnr=main(args)
        # print('read_lines',read_lines)
        # print(len(read_lines))
        # parameter=read_lines[1].strip('\n').split(',')
        # print('parameter',parameter)
        # print('parameter[1]',float(parameter[1]))
        sheet.cell(row=m+1, column=1).value = args.code
        sheet.cell(row=m+1, column=2).value = args.interp
        sheet.cell(row=m+1, column=3).value = args.max_epochs
        sheet.cell(row=m+1, column=4).value = args.batch_size
        sheet.cell(row=m+1, column=5).value = str(args.cfa_size)
        sheet.cell(row=m+1, column=6).value = 'AdamW'
        sheet.cell(row=m+1, column=7).value = float(args.slr)
        sheet.cell(row=m+1, column=8).value = float(args.mlr)
        sheet.cell(row=m+1, column=9).value = args.alpha
        sheet.cell(row=m+1, column=10).value = args.w_zhi
        sheet.cell(row=m+1, column=11).value = args.gaussian_weight_size
        sheet.cell(row=m+1, column=12).value = best_rgb_psnr
        sheet.cell(row=m+1, column=13).value = '无'
        wb.save(save_root_xlsx)
        print("数据写入成功!")
        print('11111')
    fo_main.close()
    # exit()
    # """
